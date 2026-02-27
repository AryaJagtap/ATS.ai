import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def generate_excel_bytes(results: list) -> bytes:
    df = pd.DataFrame(results)

    # Force Exact Required Columns and Ordering
    columns_order = [
        "Candidate Name", "Phone Number", "Email", "Status", "ATS Score",
        "Resume Summary", "Missing Requirements", "Job Description Summary",
        "Target Job Role", "Best Fit Role", "Resume Link", "Photo Link", "Recommendation"
    ]

    # Fill any theoretically missing columns safely
    for col in columns_order:
        if col not in df.columns:
            df[col] = "Not Found"

    # Filter and reorder
    df = df[columns_order]

    # Sort by Score
    if "ATS Score" in df.columns:
        df = df.sort_values("ATS Score", ascending=False).reset_index(drop=True)

    # Insert Serial Number at the very front
    df.insert(0, "Serial Number", range(1, len(df) + 1))

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ATS Results")
        ws = writer.sheets["ATS Results"]

        # Style Headers
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Apply Color Logic to 'ATS Score' Column
        score_col = next((i for i, col in enumerate(df.columns, 1) if col == "ATS Score"), None)
        if score_col:
            green = PatternFill("solid", fgColor="C6EFCE")
            yellow = PatternFill("solid", fgColor="FFEB9C")
            red = PatternFill("solid", fgColor="FFC7CE")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[score_col - 1]
                try:
                    val = float(cell.value)
                    if val >= 70: cell.fill = green
                    elif val >= 50: cell.fill = yellow
                    else: cell.fill = red
                except (TypeError, ValueError):
                    pass

        # Adjust Column Widths
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) > 0 else 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    return output.getvalue()